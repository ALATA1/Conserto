from fastapi import FastAPI, Request,  Body, Form, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm
from openpyxl import Workbook
from fastapi.staticfiles import StaticFiles
from app.core.security import (
    verify_password,
    create_access_token,
    has_permission
)

from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
import copy

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO
from collections import Counter

import os
import tempfile
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from typing import List, Union, Optional
from fastapi.responses import JSONResponse
from jose import jwt, JWTError
from datetime import datetime
import json
from app.database.database import engine
from app.database.base import Base
from app.core.middleware import AuditMiddleware


from starlette.middleware.sessions import SessionMiddleware
from openpyxl.styles import Font, PatternFill, Alignment
from app.data.users import users_db





# =====================
# INIT APP
# =====================
app = FastAPI()

app.mount(
    "/static",
    StaticFiles(directory="static"),
    name="static"
)

app.add_middleware(SessionMiddleware, secret_key="...")

Base.metadata.create_all(bind=engine)

app.add_middleware(AuditMiddleware)

SECRET_KEY = "conserto_secret_key"
ALGORITHM = "HS256"

collaborateurs = []
audit_logs = []

app.add_middleware(
    SessionMiddleware,
    secret_key="conserto-secret-key"
)



# ============
# BASE UNIQUE
# ============

def render_page(title, content):
    return f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <title>{title}</title>
        <link rel="stylesheet"
        href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
        <!-- <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet"> -->
    </head>

    <body class="bg-light">

        <div class="container mt-4">
            {content}
        </div>

    </body>
    </html>
    """



# ========================
# AUTH - LECTURE DU TOKEN
# ========================

# def get_current_user(request: Request):
#     token = request.cookies.get("access_token")

#     if not token:
#         return None

#     try:
#         payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
#         username = payload.get("sub")
#         return users_db.get(username)
#     except JWTError:
#         return None

SECRET_KEY = "conserto_secret_key"
ALGORITHM = "HS256"

def get_current_user(request: Request):

    token = request.cookies.get("access_token")

    if not token:
        return None

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])

        return {
            "username": payload.get("sub"),
            "role": payload.get("role")
        }

    except:
        return None
    

# =======================================
# LOGIN - FONCTION UTILISATEUR CONNECTE 
# =======================================

# @app.post("/login")
# def login(request: Request, form_data: OAuth2PasswordRequestForm = Depends()):

#     user = users_db.get(form_data.username)

#     if not user:
#         raise HTTPException(status_code=401, detail="Utilisateur incorrect")

#     if not verify_password(form_data.password, user["hashed_password"]):
#         raise HTTPException(status_code=401, detail="Mot de passe incorrect")

#     token = create_access_token({
#         "sub": form_data.username,
#         "role": user["role"]
#     })

#     response = RedirectResponse(url="/", status_code=303)
#     response.set_cookie("access_token", token, httponly=True)

#     return response

@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...)):

    user = users_db.get(username)

    if not user:
        raise HTTPException(status_code=401, detail="Utilisateur incorrect")

    if not verify_password(password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")

    token = create_access_token({
        "sub": username,
        "role": user["role"]
    })

    response = RedirectResponse(url="/", status_code=303)
    response.set_cookie(
        key="access_token",
        value=token,
        httponly=True
    )

    return response




# =========================
# HELPERS (FILTRES / UTILS)
# =========================

def filter_collaborateurs(
    data,
    agence=None,
    competence=None,
    profil=None,
    search=None,
    niveau=None
):

    data = copy.deepcopy(data)

    if agence:
        data = [c for c in data if c["agence"] == agence]

    if competence:
        data = [
            c for c in data
            if any(comp["nom"] == competence for comp in get_safe_competences(c["competence"]))
        ]

    if profil:
        data = [c for c in data if c["profil"] == profil]

    if search:
        data = [
            c for c in data
            if search.lower() in c["nom"].lower()
            or search.lower() in c["prenom"].lower()
        ]

    if niveau:
        niveau_max = int(niveau)

        data = [
            c for c in data
            if any(
                comp.get("niveau") is None
                or int(comp.get("niveau", 0)) <= niveau_max
                for comp in get_safe_competences(c["competence"])
            )
        ]

    return data


# =====================
# FONCTION SAFE
# =====================

def get_safe_competences(value):
    """
    Normalise toujours vers une liste de dictionnaires.
    """
    if not value:
        return []

    if isinstance(value, dict):
        return [value]

    if isinstance(value, list):
        result = []
        for v in value:
            if isinstance(v, dict):
                result.append(v)
            elif isinstance(v, str):
                result.append({
                    "nom": v,
                    "niveau": 0,
                    "niveau_attendu": 0,
                    "appetence": 0
                })
        return result

    if isinstance(value, str):
        return [{
            "nom": value,
            "niveau": 0,
            "niveau_attendu": 0,
            "appetence": 0
        }]

    return []


# =========================
# PROFILS
# =========================
PROFILS = [
    "Agile master",
    "Architecte data",
    "Architecte technique ",
    "Business Analyst",
    "Coach agile",
    "Concepteur Dev TU Android",
    "Concepteur Dev TU Java",
    "Concepteur fonctionnel",
    "Chef de projet",
    "Développeur",
    "Testeur fonctionnel",
    "Testeur Automaticien",
    "Data engineer",
    "Data analyst",
    "Devops",
    "Ing. Système Linux",
    "Pilotage et organisation (PMO)",
    "Product owner (PO)",
    "Sécurité et réseau"
]

# =========================
# COMPÉTENCES
# =========================

COMPETENCES = {

    # -------------------------
    # BASES DE DONNÉES
    # -------------------------

    "Bases de données": [
        "Oracle",
        "Oracle SQL Dev",
        "PG Admin",
        "PostgreSQL",
        "SQL Server"
    ],

    # -------------------------
    # DEVOPS / INFRA
    # -------------------------
    "Devops/Infra": [
        "Apache", "Docker", "Docker Compose",
        "Docker Swarm", "Git", "Gitlab",
        "GitlabCI", "Helm", "Kubernetes",
        "kubctl", "Linux Ubuntu", "Nginx",
        "Scripts Shell", "Tomcat", "SonarQube"
    ],

    # -------------------------
    # FRONTEND / MOBILE
    # -------------------------
    "Frontend & Mobile": [
        "Android",
        "Cordova"
    ],

    # -------------------------
    # FRAMEWORKS
    # -------------------------
    "Frameworks": [
        "Struts",
        "Spring",
        "Framework SPRING",
        "Framework ALT (Java spécifique)",
        "Framework SDK (.NET)",
        "Angular 17", 
        "Angular 19",
        "ROA",
        "ALF"
    ],

    # -------------------------
    # FONCTIONNEL / MÉTIER
    # -------------------------

    "Fonctionnel métier": [
        "SMARTEO",
        "FACTEO",
        "Multicanal",
        "GRC",
        "Gestion des opérations",
        "Administration des ventes",
        "Distribution",
        "Échange producteurs",
        "Dépôt et retrait",
        "Banque",
        "Fonction support",
        "Production transverse",
        "Nouveau domaine"
    ],

    # -------------------------
    # LANGAGES / DÉVELOPPEMENT
    # -------------------------
    "Langages": [
        "Apache Nifi", "CSS", "C#", "HTML",
        "Hibernate", "Java 8", "Java 21", "Java 25", "Java / J2EE"
        "Jasper report", "Javascript", "Maven 3", "NPM/NodeJS",
        "Open Feign", "RGAA", "Spring", "Spring Batch",
        "SCSS", "Typescript 5"
    ],

    # -------------------------
    # MIDDLEWARE / SERVEURS D'APPLICATION
    # -------------------------
    "Middleware / Serveurs": [
        "Websphere",
        "JBoss",
        "IIS"
    ],

    # -------------------------
    # OUTILS / IDE
    # -------------------------
    "Outils & IDE": [
        "Eclipse",
        "Visual Studio",
        "RAD / RAS"
    ],
    
    
    # -----------
    # PROTOCOLES
    # -----------
    "Protocole": ["JSON", "REST", "SOAP", "XML"],

    # ----------------------------------
    # PILOTAGE / QUALITÉ / GOUVERNANCE
    # ----------------------------------
    "Pilotage & Qualité": [
        "Pilotage",
        "Qualité / Process",
        "Pilotage et Décisionnel"
    ],

        # -----
    # PERF
    # -----
    "Perf": ["Gatling", "Jmeter"],

        
    # -------------------------
    # INFRA / SYSTÈMES
    # -------------------------
    "Systèmes & Plateformes": [
        "AIX",
        "Mainframe",
        "Windows",
        "Baradoz sur ITaaS",
        "STMC",
        "STOM"
    ],


    # -------------------------------------
    # TESTEUR (FONCTIONNEL) / AUTOMATICIEN
    # -------------------------------------
    "Tests": [
        "ArchUnit", "Cypress", "Castle Mock",
        "JUnit", "Mockito", "MockServer",
        "Postman", "Playwright",
        "Robotframework", "SoapUi",
        "Jasmine/Karma"
    ]


}

AGENCES = [
    "Bordeaux",
    "Lyon",
    "Montpellier",
    "Nantes",
    "Niort",
    "Paris",
    "Rennes",
    "Toulouse"
]

# =========================
# LEGENDES TOOLTIP
# =========================
# LEG_NIVEAU = """
# N/A = Non applicable
# 0 = Pas de notion
# 1 = Connaissances
# 2 = Travail supervisé
# 3 = Autonomie limitée
# 4 = Maîtrise et autonomie totale
# 5 = Maîtrise et enseignement
# """

# LEG_APPETENCE = """
# 0 = Aversion
# 1 = Peu d'intérêt
# 2 = Curiosité
# 3 = Intéressé
# 4 = Très intéressé
# 5 = Vital / 5 = Essentiel / 5 = Prioritaire
# """

# LEG_ATTENDU = """
# N/A = Non applicable
# De [0-1] = Débutant
# De [2-3] = Intermédiaire
# De [4-5] = Expert
# """

LEG_NIVEAU = """
N/A = Non applicable
0 = Aucune notion
1 = Connaissances théoriques
2 = Mise en pratique sous supervision
3 = Autonomie opérationnelle
4 = Maîtrise complète
5 = Expertise et capacité à enseigner
"""

LEG_APPETENCE = """
0 = Aversion
1 = Peu d'intérêt
2 = Curiosité
3 = Intérêt marqué
4 = Très intéressé
5 = Essentiel
"""

LEG_ATTENDU = """
N/A = Non applicable
[0-1] = Junior
[2-3] = Intermédiaire
4 = Confirmé
5 = Expert
"""


# =========================
# DATA
# =========================

# collaborateurs = [
#     {
#         "id": 1,
#         "nom": "ALATA",
#         "prenom": "Ibrahima",
#         "profil": "Testeur Automaticien",
#         "agence": "Niort",
#         "competence": ["Robotframework", "Playwright"],
#         "niveau": 4,
#         "niveau_attendu": 5,   
#         "appetence": 5
#     },
#     {
#         "id": 2,
#         "nom": "AVIGNON",
#         "prenom": "Martin",
#         "profil": "Développeur",
#         "agence": "Lyon",
#         "competence": ["PostgreSQL"],
#         "niveau": 3,
#         "niveau_attendu": 5,   
#         "appetence": 4
#     }
# ]


collaborateurs = [

    {
        "id": 1,
        "nom": "ALATA",
        "prenom": "Ibrahima",
        "profil": "Testeur Automaticien",
        "agence": "Niort",

        "competence": [

            {
                "nom": "Robotframework",
                "niveau": 4,
                "niveau_attendu": 5,
                "appetence": 5
            },

            {
                "nom": "Playwright",
                "niveau": 4,
                "niveau_attendu": 5,
                "appetence": 5
            },

            {
                "nom": "PostgreSQL",
                "niveau": 3,
                "niveau_attendu": 5,
                "appetence": 4
            }

        ]
    },

    {
        "id": 2,
        "nom": "AVIGON",
        "prenom": "Laurie",
        "profil": "Agile master",
        "agence": "Bordeaux",

        "competence": [

            {
                "nom": "Pilotage",
                "niveau": 3,
                "niveau_attendu": 5,
                "appetence": 5
            },

            {
                "nom": "Git",
                "niveau": 4,
                "niveau_attendu": 5,
                "appetence": 5
            }

        ]
    },

    {
        "id": 3,
        "nom": "MONDON",
        "prenom": "Romain",
        "profil": "Devops",
        "agence": "Niort",

        "competence": [

            {
                "nom": "Qualité / Process",
                "niveau": 2,
                "niveau_attendu": 5,
                "appetence": 5
            },

            {
                "nom": "Oracle",
                "niveau": 4,
                "niveau_attendu": 5,
                "appetence": 5
            }

        ]
    },

    {
        "id": 4,
        "nom": "GRIS",
        "prenom": "Philippe",
        "profil": "Développeur ",
        "agence": "Rennes",

        "competence": [

            {
                "nom": "Angular 19",
                "niveau": 3,
                "niveau_attendu": 5,
                "appetence": 5
            },

            {
                "nom": "Java 25",
                "niveau": 1,
                "niveau_attendu": 5,
                "appetence": 5
            }

        ]
    },

    {
        "id": 5,
        "nom": "REGNAULT",
        "prenom": "Justine",
        "profil": "Business Analyst ",
        "agence": "Paris",

        "competence": [

            {
                "nom": "Oracle",
                "niveau": 3,
                "niveau_attendu": 5,
                "appetence": 5
            },

            {
                "nom": "Cypress",
                "niveau": 1,
                "niveau_attendu": 5,
                "appetence": 5
            }

        ]
    }
]

# =========================
# AUDIT LOG
# =========================
audit_logs = []

def log_action(user, action, details):
    audit_logs.append({
        "date": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "user": user["username"],
        "action": action,
        "details": details
    })

# ============================
# ACTIONS CODE COULEUR NIVEAU
# ============================

def format_niveau(n):
    try:
        n = int(n)
    except:
        return '<span class="niveau-rouge" title="Valeur invalide">?</span>'

    if n <= 1:
        label = "Débutant"
        cls = "niveau-rouge"
    elif n <= 3:
        label = "Intermédiaire"
        cls = "niveau-orange"
    elif n <= 5:
        label = "Expert"
        cls = "niveau-vert"
    else:
        return f'<span class="niveau-rouge" title="Invalide">{n}</span>'

    return f'<span class="{cls}" title="{label}">{n}</span>'
    

def format_attendu(n):
    try:
        n = int(n)
    except:
        return '<span class="niveau-rouge" title="Valeur invalide">?</span>'

    if n <= 1:
        label = "Débutant"
        cls = "niveau-rouge"
    elif n <= 3:
        label = "Intermédiaire"
        cls = "niveau-orange"
    elif n <= 5:
        label = "Expert"
        cls = "niveau-vert"
    else:
        return f'<span class="niveau-rouge" title="Invalide">{n}</span>'

    return f'<span class="{cls}" title="{label}">{n}</span>'


# =========================
# HELPERS
# =========================
def normalize_data():
    for c in collaborateurs:
        if isinstance(c["competence"], str):
            c["competence"] = [c["competence"]]

normalize_data()


def normalize_competence(value):
    if not value:
        return []

    if isinstance(value, list):
        return value

    return [value]


def clamp(v):
    return max(0, min(5, int(v)))



# ============================================
# AJOUT DE LA ROUTE LOGIN BACK END - SWAGGER
# ============================================

@app.post("/login")
def login(request: Request, form_data: OAuth2PasswordRequestForm = Depends()):

    user = users_db.get(form_data.username)

    if not user:
        raise HTTPException(status_code=401, detail="Utilisateur incorrect")

    if not verify_password(form_data.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")

    token = create_access_token({
        "sub": form_data.username,
        "role": user["role"]
    })

    response = RedirectResponse(url="/", status_code=303)

    #  IMPORTANT : on stocke le token dans un cookie
    response.set_cookie(
        key="access_token",
        value=token,
        httponly=True
    )

    return response


# =========================================
# AJOUTER LOGIN AUTHENTIFICATION FRONT END
# =========================================

@app.get("/login", response_class=HTMLResponse)
def login_page():
    return """
    <html>
    <head>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        <title>Connexion</title>
    </head>

    <body class="bg-light">

    <div class="container mt-5" style="max-width:400px;">

        <h3 class="mb-3">Connexion</h3>

        <form method="post" action="/login">

            <input class="form-control mb-2" name="username" placeholder="Username" required>

            <input class="form-control mb-3" type="password" name="password" placeholder="Password" required>

            <button class="btn btn-primary w-100">Se connecter</button>

        </form>

    </div>

    </body>
    </html>
    """


# # =========================================
# # AJOUTER LOGIN AUTHENTIFICATION FRONT END
# # =========================================

# @app.get("/login", response_class=HTMLResponse)
# def login_page():

#     content = """
#     <div class="container mt-5" style="max-width:400px;">
#         <h2 class="mb-4">Connexion</h2>

#         <form method="post" action="/login">
#             <input class="form-control mb-2" name="username" placeholder="Username" required>
#             <input class="form-control mb-3" type="password" name="password" placeholder="Password" required>
#             <button class="btn btn-primary w-100">Se connecter</button>
#         </form>
#     </div>
#     """

#     return render_page("Login", content)

# @app.get("/login", response_class=HTMLResponse)
# def login_page():
#     return """
#     <html>
#     <head>
#         <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
#         <title>Connexion</title>
#     </head>

#     <body class="bg-light">

#     <div class="container mt-5" style="max-width:400px;">

#         <h3 class="mb-3">Connexion</h3>

#         <form method="post" action="/login">

#             <input class="form-control mb-2" name="username" placeholder="Username" required>

#             <input class="form-control mb-3" type="password" name="password" placeholder="Password" required>

#             <button class="btn btn-primary w-100">Se connecter</button>

#         </form>

#     </div>

#     </body>
#     </html>
#     """

# =========================
# AJOUTER LA ROUTE LOGOUT
# =========================
@app.get("/logout")
def logout():
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie("access_token")
    return response



# =========================
# ROUTES FASTAPI
# =========================

@app.post("/add_competence/{id}")
def add_competence(
    id: int,
    competence: List[str] = Form(...),
    niveau: int = Form(...),
    niveau_attendu: int = Form(...),
    appetence: int = Form(...)
):

    c = next((x for x in collaborateurs if x["id"] == id), None)

    if not c:
        return RedirectResponse("/")

    c["competence"] = normalize_competence(c["competence"])

    for comp_name in competence:

        c["competence"].append({
            "nom": comp_name,
            "niveau": clamp(niveau),
            "niveau_attendu": clamp(niveau_attendu),
            "appetence": clamp(appetence)
        })

    return RedirectResponse(f"/edit/{id}", status_code=303)





# =========================
# HOME
# =========================
    
@app.get("/", response_class=HTMLResponse)
def home(
    request: Request,
    collaborateur: str = None,
    agence: str = None,
    competence: str = None,
    search: str = None,
    profil: str = None,
    niveau: str = None,
    error: str = None
):   

    # data = collaborateurs
    data = copy.deepcopy(collaborateurs)

    current_user = get_current_user(request)

    if not current_user:
        return RedirectResponse("/login", status_code=303)
    
    username = current_user["username"]
    role = current_user["role"]

    role_label = {
        "ADMIN": "👑 Administrateur",
        "RH": "👥 Ressources Humaines",
        "MANAGER": "📊 Manager",
        "UTILISATEUR": "👤 Collaborateur"
    }.get(role, role)

    role_badge = {
        "ADMIN": "danger",
        "RH": "primary",
        "MANAGER": "warning",
        "UTILISATEUR": "secondary"
    }.get(role, "secondary")

    can_edit = has_permission(role, "edit")
    can_delete = has_permission(role, "delete")
    can_export = has_permission(role, "export")

    print(role)
    print(username)
    print(can_edit)
    print(can_delete)
    print(can_export)

    if not current_user:
        return RedirectResponse("/login", status_code=303)

    is_logged = True

    if not current_user:
        return RedirectResponse(
            url="/login",
            status_code=303
        )

    if agence:
        data = [c for c in data if c["agence"] == agence]

    if competence:
        data = [
            c for c in data
            if any(
                comp["nom"] == competence
                for comp in get_safe_competences(c["competence"])
            )
        ]
        
    if niveau:
    
        niveau_max = int(niveau)

        data = [
            c for c in data
            if any(
                (
                    comp.get("niveau") is None
                    or int(comp.get("niveau", 0)) <= niveau_max
                )
                for comp in get_safe_competences(c["competence"])
            )
        ]
    
    if profil:
        data = [c for c in data if c["profil"] == profil]

    if search:
        data = [
            c for c in data
            if search.lower() in c["nom"].lower()
            or search.lower() in c["prenom"].lower()
        ]

    
    

    # if niveau:
    
    #     niveau_max = int(niveau)

    #     data = [
    #         c for c in data
    #         if any(
    #             (
    #                 comp.get("niveau") == "N/A"
    #                 or int(comp.get("niveau", 0)) <= niveau_max
    #             )
    #             for comp in get_safe_competences(c["competence"])
    #         )
    #     ]


    if search:
        data = [
            c for c in data
            if search.lower() in c["nom"].lower()
            or search.lower() in c["prenom"].lower()
        ]

    if profil:
        data = [c for c in data if c["profil"] == profil]

    # total = sum(len(c["competence"]) for c in data)
    total = len(data)

    # avg = round(
    #     sum(
    #         comp.get("niveau", 0)
    #         for c in data
    #         for comp in normalize_competence(c.get("competence", []))
    #     ) / total,
    #     1
    # ) if total else 0

    if collaborateur:
        data = [c for c in data if c["profil"] == collaborateur]

    total = len(data)



    all_competences = [
        comp
        for c in data
        for comp in get_safe_competences(c["competence"])
    ]

    avg = round(
        sum(comp["niveau"] for comp in all_competences)
        / len(all_competences),
        1
    ) if all_competences else 0




    html = f"""
    <html>

    <head>

        <title>Skills Matrix</title>

        <link
            href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css"
            rel="stylesheet"
        >

        <style>

            body {{
                background:#f8f9fa;
            }}

            .skill-badge {{
                cursor:pointer;
                font-size:12px;
            }}

            .table td {{
                vertical-align:middle;
            }}

            .niveau-rouge {{
            color: #dc3545;
            font-weight: 600;
        }}

        .niveau-orange {{
            color: #fd7e14;
            font-weight: 600;
        }}

        .niveau-vert {{
            color: #28a745;
            font-weight: 600;
        }}
        </style>

    </head>

    <body>
    <script>
        const collaborateurs = {json.dumps(data)};
    </script>


    <div class="container mt-4">

        <div class="d-flex justify-content-between align-items-center mb-4">

            <h2>Skills Matrix</h2>

            {
                f'''
                <div class="d-flex align-items-center gap-3">
                    <!--
                    <img
                        src="https://ui-avatars.com/api/?name={username}&background=0D8ABC&color=fff"
                        class="rounded-circle border"
                        width="45"
                        height="45"
                    >
                    -->
                    <img
                        src="/static/photos/test.png"
                        width="60"
                        height="60"
                        class="rounded-circle border shadow-sm"
                    >

                    <div>

                        <div class="fw-bold">
                            {username}
                        </div>

                        <div class="d-flex align-items-center gap-2 mt-1">

                            <span class="badge bg-{role_badge}"
                            style="min-width:190px;">
                                {role_label}
                            </span>

                            <a href="/logout"
                            class="badge bg-dark text-decoration-none px-3 py-2"
                            style="min-width:190px;">

                                <i class="bi bi-box-arrow-right me-1"></i>
                                Déconnexion

                            </a>

                        </div>

                    </div>

                </div>

                    
                
                '''
                if is_logged else
                '''
                <a href="/login" class="btn btn-primary">
                    Connexion
                </a>
                '''
            }

        </div>

        




        <div class="alert alert-info fw-bold ">
            Échelle de notation : Survoler les champs niveau et appétence pour voir les légendes.
        </div>

        <!--   BLOCS Total collabs et Niveau moyen des compétences -->

        <div class="row mb-3">
            <div class="col card p-3 m-2 shadow-sm border-start border-primary border-4">
                <span style="font-size:18px;">
                    👥 <strong>Total collaborateurs :</strong>
                    <span style="
                        color:#0d6efd;
                        font-size:22px;
                        font-weight:bold;">
                        {total}
                    </span>

                    <span
                        title="Nombre de collaborateurs affichés après application des filtres sélectionnés."
                        style="
                            cursor:help;
                            margin-left:5px;">
                        ℹ️
                    </span>

                </span>
            </div>

            <div class="col card p-3 m-2 shadow-sm border-start border-success border-4">
                <span style="font-size:18px;">
                    📊 <strong>Niveau moyen des compétences :</strong>
                    <span style="
                        color:#198754;
                        font-size:22px;
                        font-weight:bold;">
                        {avg}/5
                    </span>

                    <span
                        title="Calcul : somme des niveaux de toutes les compétences affichées divisée par le nombre total de compétences affichées."
                        style="
                            cursor:help;
                            margin-left:5px;">
                        ℹ️
                    </span>
                </span>
            </div>
        </div>

        <!-- EXPORT -->
        <div class="mb-3">

            <a href="/export/excel?agence={agence or ''}&competence={competence or ''}&profil={profil or ''}&search={search or ''}"
               class="btn btn-dark">
                Export Excel
            </a>

            <a href="/export/pdf?agence={agence or ''}&competence={competence or ''}&profil={profil or ''}&search={search or ''}"
                class="btn btn-danger">
                 Export PDF
            </a>

        </div>

        <!-- ADD -->
        <form class="card p-3 mb-4" action="/add" method="post">

            <div class="row">
                <div class="col"><input class="form-control" name="nom" placeholder="Nom*" required></div>
                <div class="col"><input class="form-control" name="prenom" placeholder="Prénom*" required></div>

                <div class="col">
                    <select class="form-control" name="profil" required>
                        <option value="">Profil*</option>
                        {"".join([f'<option value="{p}">{p}</option>' for p in PROFILS])}
                    </select>
                </div>

                <div class="col">
                    <select class="form-control" name="agence" required>
                        <option value="">Agence*</option>
                        {"".join([f'<option value="{a}">{a}</option>' for a in AGENCES])}
                    </select>
                </div>
            </div>

            <div class="row mt-2">

                <!-- COMPETENCE -->
                <div class="col">

                    <div class="dropdown w-100">

                        <button
                            class="btn btn-outline-secondary dropdown-toggle w-100 text-start d-flex align-items-center"
                            type="button"
                            data-bs-toggle="dropdown"
                            id="competenceBtn"
                            style="height:38px;"
                        >
                            Compétences*
                        </button>

                        <div class="dropdown-menu p-3 w-100"
                            style="max-height:250px; overflow:auto;">

                            {''.join([
                                f"<div class='fw-bold mt-2'>{g}</div>" +
                                ''.join([
                                    f"""
                                    <div class="form-check">

                                        <input
                                            class="form-check-input competence-check"
                                            type="checkbox"
                                            name="competence"
                                            value="{s}"
                                            onchange="validateCompetence()"
                                        >

                                        <label class="form-check-label">
                                            {s}
                                        </label>

                                    </div>
                                    """
                                    for s in skills
                                ])
                                for g, skills in COMPETENCES.items()
                            ])}

                        </div>

                    </div>

                    <!-- VALIDATEUR HTML5 -->
                    <input
                        type="text"
                        id="competence-validator"
                        required
                        style="
                            position:absolute;
                            opacity:0;
                            height:0;
                            width:0;
                            pointer-events:none;
                        "
                    >

                </div>

                <script>

                    function validateCompetence() {{

                        const checked =
                            document.querySelectorAll(".competence-check:checked");

                        const validator =
                            document.getElementById("competence-validator");

                        if (checked.length > 0) {{
                            validator.value = "ok";
                        }} else {{
                            validator.value = "";
                        }}
                    }}

                    validateCompetence();

                </script>


                <!--
                <div class="col">
                    <input class="form-control" name="niveau" type="number"
                        min="0" max="5" placeholder="Niveau*"
                        title="{LEG_NIVEAU}" required>
                </div>
                -->

                <div class="col">

                    <select class="form-control"
                            name="niveau"
                            title="{LEG_NIVEAU}">

                        <option value="">
                            Niveau*
                        </option>

                        <option value="0">
                            N/A
                        </option>

                        <option value="1">
                            1
                        </option>

                        <option value="2">
                            2
                        </option>

                        <option value="3">
                            3
                        </option>

                        <option value="4">
                            4
                        </option>

                        <option value="5">
                            5
                        </option>

                    </select>

                </div>

                <div class="col">
                    <input class="form-control" name="niveau_attendu" type="number"
                        min="0" max="5" placeholder="Niveau attendu*"
                        title="{LEG_ATTENDU}" required>
                </div>
             
                <div class="col">
                    <input class="form-control" name="appetence" type="number"
                        min="0" max="5" placeholder="Appétence*"
                        title="{LEG_APPETENCE}" required>
                </div>

                <div class="col">
                    <button class="btn btn-primary w-100">Ajouter</button>
                </div>
            </div>
        </form>

        <!-- FILTERS -->
        <form class="card p-3 mb-4"
              method="get">

            <div class="row">

                <div class="col">
                    <input
                        class="form-control"
                        name="search"
                        placeholder="Nom / Prénom"
                    >
                </div>

                <div class="col">

                    <select class="form-control"
                            name="profil">

                        <option value="">
                            Profil
                        </option>

                        {"".join([
                            f'<option value="{p}">{p}</option>'
                            for p in PROFILS
                        ])}

                    </select>

                </div>

                <div class="col">

                    <select class="form-control"
                            name="agence">

                        <option value="">
                            Agence
                        </option>

                        {"".join([
                            f'<option value="{a}">{a}</option>'
                            for a in AGENCES
                        ])}

                    </select>

                </div>

                <div class="col">

                    <select class="form-control"
                            name="competence">

                        <option value="">
                            Compétence
                        </option>

                        {"".join([
                            f'''
                            <optgroup label="{g}">
                                {"".join([
                                    f'<option value="{s}">{s}</option>'
                                    for s in skills
                                ])}
                            </optgroup>
                            '''
                            for g, skills in COMPETENCES.items()
                        ])}

                    </select>

                </div>

                <div class="col">
                    <input class="form-control" name="niveau" type="number"
                        min="0" max="5" placeholder="Niveau"
                        title="{LEG_NIVEAU}" required>
                </div>

                <div class="col">
                    <button class="btn btn-success w-100">
                        Filtrer
                    </button>
                </div>

                <div class="col">
                    <a href="/"
                       class="btn btn-secondary w-100">
                        Reset
                    </a>
                </div>

            </div>

        </form>

        <!-- TABLE -->
        <!--
        <table class="table table-striped table-bordered bg-white">

            <thead class="table-dark">

                <tr>
                    <th>Nom</th>
                    <th>Prénom</th>
                    <th>Profil</th>
                    <th>Agence</th>
                    <th>Compétences</th>
                    <th>Niveau</th>
                    <th>Niveau attendu</th>
                    <th>Appétence</th>
                    <th>Actions</th>
                </tr>

            </thead>

            <tbody>
             -->

        <div style="display:flex; gap:20px;">
    
    <!-- TABLE -->
    <div style="flex:2;">
        <!--<div id="detailPanel" style="flex:1; background:white; padding:15px; border-radius:8px;">
            <div class="text-muted">
                Clique sur une ligne pour voir les détails
            </div>
        </div> -->
        <table class="table table-striped table-bordered bg-white">

            <thead class="table-dark">
                <tr>
                    <th>Nom</th>
                    <th>Prénom</th>
                    <th>Profil</th>
                    <th>Agence</th>
                    <th>Compétences</th>
                    <th>Niveau</th>
                    <th>Niveau attendu</th>
                    <th>Appétence</th>
                    <th>Actions</th>
                </tr>
            </thead>
            <script>
                const collaborateurs = {{collabs}};
            </script>
            <tbody>
           
    """
    
    
    # for c in data:

    #     competences_html = ""

    #     for comp in c["competence"]:

    #         competences_html += f"""
    #         <span
    #             class="badge bg-primary me-1 mb-1 skill-badge"
    #             onclick="openDeleteModal(
    #                 {c['id']},
    #                 '{comp}',
    #                 {len(c["competence"])}
    #             )"
    #         >
    #             {comp} ✕
    #         </span>
    #         """

    #     html += f"""
    #     <tr style="cursor:pointer"
    #     onclick="showDetails({c['id']})">

    #         <td>{c['nom']}</td>
    #         <td>{c['prenom']}</td>
    #         <td>{c['profil']}</td>
    #         <td>{c['agence']}</td>

    #         <td>
    #             {
    #                 "".join([

    #                     f"""
    #                     <span
    #                         class="badge bg-primary me-1 mb-1 skill-badge"
    #                         onclick="openDeleteModal(
    #                             {c['id']},
    #                             '{comp}',
    #                             {len(c["competence"])}
    #                         )"
    #                     >
    #                         {comp} ✕
    #                     </span>
    #                     """

    #                     for comp in c["competence"]

    #                 ])
    #             }
    #         </td>
    #         <td>{format_niveau(c['niveau'])}</td>
            
    #         <td>{c.get('niveau_attendu', '')}</td>
    #         <td>{c['appetence']}</td>

    #         <td onclick="event.stopPropagation()">

    #             <a href="/edit/{c['id']}"
    #             class="btn btn-warning btn-sm">
    #                 ✏️
    #             </a>

    #             <button
    #                 class="btn btn-danger btn-sm"
    #                 onclick="openDeleteCollaborateurModal(
    #                     {c['id']},
    #                     '{c['nom']}',
    #                     '{c['prenom']}'
    #                 )"
    #             >
    #                 🗑
    #             </button>

    #         </td>

    #     </tr>
    #     """

    for c in data:
    
        nb_competences = len(c["competence"])

        for index, comp in enumerate(c["competence"]):

            html += '<tr style="cursor:pointer">'

            if index == 0:

                html += f"""
                    <td rowspan="{nb_competences}">
                        {c['nom']}
                    </td>

                    <td rowspan="{nb_competences}">
                        {c['prenom']}
                    </td>

                    <td rowspan="{nb_competences}">
                        {c['profil']}
                    </td>

                    <td rowspan="{nb_competences}">
                        {c['agence']}
                    </td>
                """

            html += f"""

                <td>

                    <span
                        class="badge bg-primary me-1 mb-1 skill-badge"
                        onclick="openDeleteModal(
                            {c['id']},
                            '{comp['nom']}',
                            {nb_competences}
                        )"
                    >
                        {comp['nom']} ✕
                    </span>

                </td>

                <td>
                    {format_niveau(comp['niveau'])}
                </td>

                <td>
                    {format_attendu(comp['niveau_attendu'])}
                </td>

                <td>
                    {comp['appetence']}
                </td>

                <td onclick="event.stopPropagation()">

                    <a
                        href="/edit/{c['id']}"
                        class="btn btn-warning btn-sm"
                    >
                        ✏️
                    </a>

                    <button
                        class="btn btn-danger btn-sm"
                        onclick="openDeleteCollaborateurModal(
                            {c['id']},
                            '{c['nom']}',
                            '{c['prenom']}'
                        )"
                    >
                        🗑
                    </button>

                </td>

            """

            html += "</tr>"
    
    html += """
            </tbody>

        </table>

    </div>

    <!-- MODAL DELETE COMPETENCE -->
    <div class="modal fade"
         id="deleteCompetenceModal"
         tabindex="-1">

        <div class="modal-dialog">

            <div class="modal-content">

                <div class="modal-header">

                    <h5 class="modal-title"
                        id="deleteModalTitle">
                    </h5>

                    <button type="button"
                            class="btn-close"
                            data-bs-dismiss="modal">
                    </button>

                </div>

                <div class="modal-body"
                     id="deleteModalBody">
                </div>

                <div class="modal-footer">

                    <button type="button"
                            class="btn btn-secondary"
                            data-bs-dismiss="modal">
                        Annuler
                    </button>

                    <a id="confirmDeleteBtn"
                       href="#"
                       class="btn btn-danger">
                        Supprimer
                    </a>

                </div>

            </div>

        </div>

    </div>

    <!-- MODAL DELETE COLLABORATEUR -->
    <div class="modal fade"
         id="deleteCollaborateurModal"
         tabindex="-1">

        <div class="modal-dialog">

            <div class="modal-content">

                <div class="modal-header">

                    <h5 class="modal-title">
                        Supprimer le collaborateur
                    </h5>

                    <button type="button"
                            class="btn-close"
                            data-bs-dismiss="modal">
                    </button>

                </div>

                <div class="modal-body"
                     id="deleteCollaborateurBody">
                </div>

                <div class="modal-footer">

                    <button type="button"
                            class="btn btn-secondary"
                            data-bs-dismiss="modal">
                        Annuler
                    </button>

                    <a id="confirmDeleteCollaborateurBtn"
                       href="#"
                       class="btn btn-danger">
                        Supprimer
                    </a>

                </div>

            </div>

        </div>

    </div>
    """

    if error == "last_competence":

        html += """
        <div class="modal fade show"
             id="infoModal"
             tabindex="-1"
             style="display:block; background:rgba(0,0,0,0.5);">

            <div class="modal-dialog">

                <div class="modal-content">

                    <div class="modal-header">

                        <h5 class="modal-title">
                            Suppression impossible
                        </h5>

                    </div>

                    <div class="modal-body">

                        Le collaborateur doit conserver
                        au moins une compétence.

                    </div>

                    <div class="modal-footer">

                        <a href="/"
                           class="btn btn-primary">
                            OK
                        </a>

                    </div>

                </div>

            </div>

        </div>
        """

    html += """
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    
    <script>

        // =========================
        // DELETE COMPETENCE
        // =========================
        function openDeleteModal(id, competence, nbCompetences) {

            const title =
                document.getElementById("deleteModalTitle");

            const body =
                document.getElementById("deleteModalBody");

            const confirmBtn =
                document.getElementById("confirmDeleteBtn");

            if (parseInt(nbCompetences) <= 1) {

                title.innerText =
                    "Suppression impossible";

                body.innerHTML =
                    "Le collaborateur doit conserver " +
                    "au moins une compétence.";

                confirmBtn.style.display = "none";

            } else {

                title.innerText =
                    "Supprimer la compétence";

                body.innerHTML =
                    "Voulez-vous vraiment supprimer : " +
                    "<b>" + competence + "</b> ?";

                confirmBtn.style.display = "inline-block";

                confirmBtn.href =
                    "/remove-competence/" +
                    id +
                    "?comp=" +
                    encodeURIComponent(competence);
            }

            const modal = new bootstrap.Modal(
                document.getElementById(
                    "deleteCompetenceModal"
                )
            );

            modal.show();
        }

        // =========================
        // DELETE COLLABORATEUR
        // =========================
        function openDeleteCollaborateurModal(
            id,
            nom,
            prenom
        ) {

            const body =
                document.getElementById(
                    "deleteCollaborateurBody"
                );

            const confirmBtn =
                document.getElementById(
                    "confirmDeleteCollaborateurBtn"
                );

            body.innerHTML =
                "Voulez-vous vraiment supprimer : <br>" +
                "<b>" +
                nom +
                " " +
                prenom +
                "</b> ?";

            confirmBtn.href =
                "/delete/" + id;

            const modal = new bootstrap.Modal(
                document.getElementById(
                    "deleteCollaborateurModal"
                )
            );

            modal.show();
        }

    </script>

        
    </body>
    </html>
    """

    return HTMLResponse(content=html)



# =========================
# ADD/AJOUT
# =========================

@app.post("/add")
def add(
    request: Request,
    nom: str = Form(...),
    prenom: str = Form(...),
    profil: str = Form(...),
    agence: str = Form(...),
    competence: Optional[Union[List[str], str]] = Form(None),
    niveau: int = Form(...),
    niveau_attendu: int = Form(...),
    appetence: int = Form(...)
):

    user = get_current_user(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    # 🔐 CONTRÔLE ROLE
    if not has_permission(user["role"], "edit"):
        raise HTTPException(
            status_code=403,
            detail="Ajout interdit"
        )

    competence = normalize_competence(competence)

    collaborateurs.append({
        "id": max([c["id"] for c in collaborateurs], default=0) + 1,
        "nom": nom,
        "prenom": prenom,
        "profil": profil,
        "agence": agence,
        "competence": [
            {
                "nom": comp,
                "niveau": clamp(niveau),
                "niveau_attendu": clamp(niveau_attendu),
                "appetence": clamp(appetence)
            }
            for comp in competence
        ]
    })

    return RedirectResponse("/", status_code=303)


# @app.post("/add")
# def add(
#     request: Request,
#     nom: str = Form(...),
#     prenom: str = Form(...),
#     profil: str = Form(...),
#     agence: str = Form(...),
#     competence: Optional[Union[List[str], str]] = Form(None),
#     niveau: int = Form(...),
#     niveau_attendu: int = Form(...),
#     appetence: int = Form(...)
# ):

#     print("DEBUG niveau_attendu =", niveau_attendu)  

#     user = get_current_user(request)

#     if not user:
#         raise HTTPException(status_code=401, detail="Non autorisé")

#     competence = normalize_competence(competence)

#     collaborateurs.append({
#         "id": max([c["id"] for c in collaborateurs], default=0) + 1,
#         "nom": nom,
#         "prenom": prenom,
#         "profil": profil,
#         "agence": agence,

#         "competence": [
#             {
#                 "nom": comp,
#                 "niveau": clamp(niveau),
#                 "niveau_attendu": clamp(niveau_attendu),
#                 "appetence": clamp(appetence)
#             }
#             for comp in competence
#         ]
#     })

#     log_action(user, "AJOUT", f"{prenom} {nom}")

#     return RedirectResponse("/", status_code=303)





# =========================
# DELETE COLLABORATEUR
# =========================
@app.get("/delete/{id}")
def delete(request: Request, id: int):

    user = get_current_user(request)

    if not has_permission(user["role"], "delete"):
        raise HTTPException(
            status_code=403,
            detail="Vous n'avez pas le droit de supprimer"
        )

    # if not user:
    #     raise HTTPException(status_code=401, detail="Non autorisé")

    global collaborateurs

    collaborateur = next((c for c in collaborateurs if c["id"] == id), None)

    collaborateurs = [
        c for c in collaborateurs
        if c["id"] != id
    ]

    if collaborateur:
        log_action(user, "SUPPRESSION", f"{collaborateur['prenom']} {collaborateur['nom']}")

    return RedirectResponse("/", status_code=303)


# =========================
# REMOVE COMPETENCE
# =========================
@app.get("/remove-competence/{id}")
def remove_competence(id: int, comp: str):

    for c in collaborateurs:

        if c["id"] == id:

            if len(c["competence"]) <= 1:
                return RedirectResponse("/?error=last_competence", status_code=303)

            c["competence"] = [
                x for x in c["competence"]
                if x["nom"] != comp
            ]
            break

    return RedirectResponse("/", status_code=303)


# ======================================
# ROUTE UNIQUE DE DÉCISION PAGES 1 et 2
# ======================================

@app.get("/edit/{id}")
def edit_router(request: Request, id: int):

    c = next((x for x in collaborateurs if x["id"] == id), None)

    if not c:
        return RedirectResponse("/")

    c = c.copy()
    c["competence"] = normalize_competence(c["competence"])

    # CAS 1 : 0 ou 1 compétence → PAGE SIMPLE
    if not c.get("competence") or len(c["competence"]) == 1:
        return edit_global_page(id)

    # CAS 2 : plusieurs compétences → PAGE LISTE
    return edit_competence_page(request,id)


# =========================
# EDIT PAGE 1 (GLOBAL)
# =========================

def edit_global_pageSSS(id: int):

    c = next((x for x in collaborateurs if x["id"] == id), None)

    if not c:
        return RedirectResponse("/")

    c = c.copy()
    c["competence"] = normalize_competence(c["competence"])

    html = f"""
    <html>
    <head>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        <title>Edit collaborateur</title>
    </head>

    <body class="bg-light">
    <div class="container mt-4">

        <h3>Modifier collaborateur</h3>

        <form method="post" action="/update/{id}" class="card p-4">

            <div class="row g-3">

                <div class="col-md-6">
                    <label>Nom</label>
                    <input class="form-control" name="nom" value="{c['nom']}">
                </div>

                <div class="col-md-6">
                    <label>Prénom</label>
                    <input class="form-control" name="prenom" value="{c['prenom']}">
                </div>

                <div class="col-md-6">
                    <label>Profil</label>
                    <input class="form-control" name="profil" value="{c['profil']}">
                </div>

                <div class="col-md-6">
                    <label>Agence</label>
                    <input class="form-control" name="agence" value="{c['agence']}">
                </div>

            </div>

            <hr>

            <h5>Compétence unique</h5>

            <input type="hidden"
                   name="competence"
                   value="{c['competence'][0]['nom'] if c.get('competence') else ''}">

            <div class="row g-3 mt-2">

                <div class="col">
                    <label>Niveau</label>
                    <input class="form-control" name="niveau"
                           value="{c['competence'][0]['niveau'] if c.get('competence') else 0}">
                </div>

                <div class="col">
                    <label>Niveau attendu</label>
                    <input class="form-control" name="niveau_attendu"
                           value="{c['competence'][0]['niveau_attendu'] if c.get('competence') else 0}">
                </div>

                <div class="col">
                    <label>Appétence</label>
                    <input class="form-control" name="appetence"
                           value="{c['competence'][0]['appetence'] if c.get('competence') else 0}">
                </div>

            </div>

            <button class="btn btn-success mt-3">Sauvegarder</button>

        </form>

    </div>
    </body>
    </html>
    """

    return HTMLResponse(html)


# =========================
# EDIT PAGE 2 (MULTI COMPÉTENCES)
# =========================
@app.get("/edit/{id}")
def edit_competence_page(request: Request, id: int):
    
    current_user = get_current_user(request)

    if not current_user:
        return RedirectResponse("/login", status_code=303)
    
    username = current_user["username"]
    role = current_user["role"]

    can_edit = has_permission(role, "edit")
    can_delete = has_permission(role, "delete")
    can_export = has_permission(role, "export")

    print(role)
    print(username)
    print(can_edit)
    print(can_delete)
    print(can_export)

    c = next((x for x in collaborateurs if x["id"] == id), None)

    if not c:
        return RedirectResponse("/")

    c = c.copy()
    c["competence"] = normalize_competence(c["competence"])

    add_button = ""

    if can_edit:
        add_button = """
        <div class="d-flex justify-content-end gap-2">
            <button class="btn btn-success" onclick="toggleAddCompetence()">
                ➕ Ajouter nouvelle compétence
            </button>

        <!-- NOUVEAU BOUTON VALIDATION -->
            <button
                type="button"
                class="btn btn-primary"
                onclick="saveAllCompetences()">
                💾 Valider tout
            </button>
        
        <!-- PAGE MODIFIER OU AJOUTER COMPETENCE - BLOC BOUTON RESET --> 
            <div class="col"> 
                <a href="/" 
                class="btn btn-secondary w-100" 
                onclick="return confirmCancel(event)"> 
                🏠 Retour Home 
                </a> 
            </div>
        </div>
        """


    html = f"""
    <html>
    <head>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        <title>Edit compétences</title>
    </head>

    <body class="bg-light">

    <div class="container mt-4">
     
        <div class="d-flex justify-content-between align-items-center mb-3">

            <!-- TITRE -->
            <h3 class="mb-0 fw-bold"
                style="
                    background: linear-gradient(90deg, #0d6efd, #20c997);
                    -webkit-background-clip: text;
                    -webkit-text-fill-color: transparent;
                    font-size: 26px;
                    display: flex;
                    align-items: center;
                    gap: 8px;
                ">
                🛠️ Modifier ou ajouter une compétence
            </h3>

            {add_button}

        </div>
        
 
            <!-- ACTIONS -->
            <div class="d-flex gap-2">

                <!-- bouton ajout -->
                {
                f'''
                <!--
                <button
                    type="button"
                    class="btn btn-success"
                    onclick="toggleAddCompetence()">

                    + Ajouter nouvelle compétence
                </button> 
                -->
                
                '''
                if can_edit else ''
                }

                

            </div>

        </div>

        
        <div class="container">
            <div id="addCompetenceForm"
                class="card p-3 mb-3"
                style="display:none;">
 
                <h7>Nouvelle compétence</h7>       

                <form method="post" action="/add_competence/{id}">

                    <div class="row mt-2">

                        <!-- COMPETENCE -->
                        <div class="col">

                            <div class="dropdown w-100">

                                <button
                                    class="btn btn-outline-secondary dropdown-toggle w-100 text-start d-flex align-items-center"
                                    type="button"
                                    data-bs-toggle="dropdown"
                                    id="competenceBtn"
                                    style="height:38px;"
                                >
                                    Compétences*
                                </button>

                                <div class="dropdown-menu p-3 w-100"
                                    style="max-height:250px; overflow:auto;">

                                    {''.join([
                                        f"<div class='fw-bold mt-2'>{g}</div>" +

                                        ''.join([

                                            f'''
                                            <div class="form-check">

                                                <input
                                                    class="form-check-input competence-check"
                                                    type="checkbox"
                                                    name="competence"
                                                    value="{s}"
                                                    onchange="validateCompetence()"
                                                >

                                                <label class="form-check-label">
                                                    {s}
                                                </label>

                                            </div>
                                            '''

                                            for s in skills

                                        ])

                                        for g, skills in COMPETENCES.items()
                                    ])}

                                </div>

                            </div>

                            <input
                                type="text"
                                id="competence-validator"
                                required
                                style="
                                    position:absolute;
                                    opacity:0;
                                    height:0;
                                    width:0;
                                    pointer-events:none;
                                "
                            >

                        </div>

                        <!-- PAGE MODIFIER OU AJOUTER COMPETENCE - BLOC NIVEAU  -->

                        <div class="col">
                            <input class="form-control" name="niveau" type="number"
                                min="0" max="5" placeholder="Niveau*"
                                title="{LEG_NIVEAU}" required>
                        </div>


                        <!-- PAGE MODIFIER OU AJOUTER COMPETENCE - BLOC NIVEAU ATTENDU -->
                        <div class="col">
                            <input class="form-control" name="niveau_attendu" type="number"
                                min="0" max="5" placeholder="Niveau attendu*"
                                title="{LEG_ATTENDU}" required>
                        </div>
                    
                        

                        <!-- PAGE MODIFIER OU AJOUTER COMPETENCE - BLOC APPETENCE -->
                        <div class="col">
                            <input class="form-control" name="appetence" type="number"
                                min="0" max="5" placeholder="Appétence*"
                                title="{LEG_APPETENCE}" required>
                        </div>

                        
                        <!-- PAGE MODIFIER OU AJOUTER COMPETENCE - BLOC BOUTON AJOUTER -->
                        <div class="col">
                            <button class="btn btn-primary w-100">
                                Ajouter
                            </button>
                        </div>

                        <!-- PAGE MODIFIER OU AJOUTER COMPETENCE - BLOC BOUTON RESET -->

                        <div class="col">
                            <button
                                type="button"
                                class="btn btn-secondary w-100"
                                onclick="document.getElementById('addCompetenceForm').style.display='none'">
                                Reset
                            </button>
                        </div>



                        

                    </div>

                </form>

            </div>

        </div>   

            

            </div>

        </div>

        <div class="row justify-content-center">

            <div class="col-md-8">

                <div class="card p-4">

    """

    # 🔥 BOUCLE COMPÉTENCES
    for comp in c["competence"]:

        delete_button = ""

        if can_delete:
            delete_button = f"""
            <a
                href="/delete_competence/{id}/{comp['nom']}"
                class="btn btn-danger"
                onclick="return confirm('Supprimer la compétence {comp["nom"]} ?')">
                Supprimer
            </a>
            """

        html += f"""
        <form method="post" action="/update/{id}/competence/{comp['nom']}" class="border p-3 mb-3">

            <h5>{comp['nom']}</h5>

            <div class="row g-2">

                <div class="col">
                    <label>Niveau</label>
                    <input class="form-control"
                           name="niveau"
                           value="{comp['niveau']}">
                </div>

                <div class="col">
                    <label>Niveau attendu</label>
                    <input class="form-control"
                           name="niveau_attendu"
                           value="{comp['niveau_attendu']}">
                </div>

                <div class="col">
                    <label>Appétence</label>
                    <input class="form-control"
                           name="appetence"
                           value="{comp['appetence']}">
                </div>

                <div class="col d-flex justify-content-end align-items-end gap-2">
                    <button class="btn btn-primary"
                        <span>Enregistrer</span>
                        
                    </button>

                    {delete_button}

                    <a
                        href="/"
                        class="btn btn-secondary"
                        onclick="return confirmCancel(event)">
                        Annuler
                    </a>

                <!--
                    <button class="btn btn-primary">
                        Enregistrer
                    </button>

                    <a
                        href="/delete_competence/{id}/{comp['nom']}"
                        class="btn btn-danger"
                        onclick="return confirm('Supprimer la compétence {comp["nom"]} ?')">
                        Supprimer
                    </a>

                    <a
                        href="/"
                        class="btn btn-secondary"
                        onclick="return confirmCancel(event)">
                        Annuler
                    </a>
                -->
                </div>

                


            </div>

        </form>
        """

    html += """
            </div>
        </div>

        <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>

        <script>

        function toggleAddCompetence() {{

            const form = document.getElementById("addCompetenceForm");

            if (form.style.display === "none") {{
                form.style.display = "block";
            }}
            else {{
                form.style.display = "none";
            }}

        }}

        </script>
        
        <script>

        function validateCompetence() {

            const checked =
                document.querySelectorAll(".competence-check:checked");

            const validator =
                document.getElementById("competence-validator");

            if (checked.length > 0) {
                validator.value = "ok";
            } else {
                validator.value = "";
            }
        }

        validateCompetence();

        </script>


        <script>
            function saveAllCompetences() {

                if (confirm("Valider toutes les modifications ?")) {

                    alert("Toutes les compétences ont été enregistrées avec succès !");

                    // option 1 : recharger la page
                    location.reload();

                    // option 2 (si tu veux plus tard) :
                    // window.location.href = "/save-all"
                }
            }
            </script>

            </body>
            </html>

        </body>
        </html>
    """

    return HTMLResponse(html)






# =========================
# EDIT PAGE OZO
# =========================

def edit_global_page(id):

    c = next((x for x in collaborateurs if x["id"] == id), None)

    if not c:
        return RedirectResponse("/")

    c["competence"] = normalize_competence(c["competence"])

    html = f"""
    <html>
    <head>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        <title>Edit</title>
        <style>
            .required-star {{
                color: red;
                margin-left: 3px;
            }}
        </style>
    </head>

    <body class="bg-light">
    <div class="container mt-4">

        <h3 style="color: #ff69b4;">
            Modifier la matrice de compétance du collaborateur
        </h3>



        <div class="alert alert-info">
            Survoler les champs pour afficher les légendes :
            Niveau (compétence) / Appétence (motivation)
        </div>

        <form class="card p-4" method="post" action="/update/{id}">

            <div class="row g-3">

                <!-- NOM -->
                <div class="col-12 col-md-6">
                    <label class="form-label">
                        Nom <span class="text-danger">*</span>
                    </label>
                    <input class="form-control" name="nom" value="{c['nom']}" required>
                </div>

                <!-- PRENOM -->
                <div class="col-12 col-md-6">
                    <label class="form-label">
                        Prénom <span class="text-danger">*</span>
                    </label>
                    <input class="form-control" name="prenom" value="{c['prenom']}" required>
                </div>

                <!-- PROFIL -->
                <div class="col-12 col-md-4">
                    <label class="form-label">
                        Profil <span class="text-danger">*</span>
                    </label>
                    <select class="form-select" name="profil" required>
                        {"".join([f'<option value="{p}" {"selected" if p==c["profil"] else ""}>{p}</option>' for p in PROFILS])}
                    </select>
                </div>

                <!-- AGENCE -->
                <div class="col-12 col-md-4">
                    <label class="form-label">
                        Agence <span class="text-danger">*</span>
                    </label>
                    <select class="form-select" name="agence" required>
                        {"".join([f'<option value="{a}" {"selected" if a==c["agence"] else ""}>{a}</option>' for a in AGENCES])}
                    </select>
                </div>

                <!-- COMPETENCE -->
                
                <div class="col-12 col-md-4">
                    <label class="form-label">
                        Compétence <span class="text-danger">*</span>
                    </label>

                   <!-- COMPETENCE VERSION LISTE DEROULANTE --> 
                   <!--
                    <select class="form-select" name="competence" required>

                        <option value="">Compétence</option>

                        {"".join([
                            f'''
                            <optgroup label="{g}">
                                {"".join([
                                    f'<option value="{s}" {"selected" if s == c["competence"] else ""}>{s}</option>'
                                    for s in skills
                                ])}
                            </optgroup>
                            '''
                            for g, skills in COMPETENCES.items()
                        ])}

                    </select>
                    -->
                    
                    <!-- COMPETENCE VERSION CASES A COCHER --> 
    
                    <div class="dropdown w-100">

                        <button
                            class="btn btn-outline-secondary dropdown-toggle w-100 text-start d-flex align-items-center"
                            type="button"
                            data-bs-toggle="dropdown"
                            id="competenceBtn"
                            style="height:38px;"
                        >
                            Sélectionner vos compétences
                        </button>

                        <div class="dropdown-menu p-3 w-100"
                            style="max-height:250px; overflow:auto;">

                            {''.join([

                                f"<div class='fw-bold mt-2'>{g}</div>" +

                                ''.join([

                                    f"""
                                    <div class="form-check">

                                        <input
                                            class="form-check-input competence-check"
                                            type="checkbox"
                                            name="competence"
                                            value="{s}"
                                            {"checked" if s in c["competence"] else ""}
                                        >

                                        <label class="form-check-label">
                                            {s}
                                        </label>

                                    </div>
                                    """

                                    for s in skills

                                ])

                                for g, skills in COMPETENCES.items()

                            ])}

                        </div>

                    </div>

                </div>

                <!-- NIVEAU -->
                <div class="col-12 col-md-4">
                    <label class="form-label">
                        Niveau <span class="text-danger">*</span>
                    </label>
                    <input class="form-control" name="niveau" type="number"
                        min="0" max="5"
                        value="{c['competence'][0]['niveau'] if c.get('competence') else 0}">
                        title="{LEG_NIVEAU}" required>
                </div>

                


                <!-- NIVEAU ATTENDU -->
                <div class="col-12 col-md-4">
                    <label class="form-label">
                        Niveau attendu <span class="text-danger">*</span>
                    </label>
                    <input class="form-control"
                        name="niveau_attendu"
                        type="number"
                        min="0"
                        max="5"
                        value="{c.get('niveau_attendu', '')}"
                        placeholder="Niveau attendu*"
                        required>
                </div>

                <!-- APPETENCE -->
                <div class="col-12 col-md-4">
                    <label class="form-label">
                        Appétence <span class="text-danger">*</span>
                    </label>
                    <input class="form-control" name="appetence" type="number"
                        min="0" max="5"
                        value="{c['competence'][0]['appetence'] if c.get('competence') else 0}">
                        title="{LEG_APPETENCE}" required>
                </div>

            </div>

            <!-- ACTIONS -->
            <div class="mt-4 d-flex gap-2">

                <!-- <button type="button" class="btn btn-success" onclick="confirmSave(event)">  
                <button class="btn btn-success"> -->

                <button type="button" id="saveBtn" class="btn btn-success" disabled onclick="confirmSave(event)">
                    Sauvegarder
                </button>

                <a href="/" class="btn btn-secondary" onclick="confirmCancel(event)">
                    Annuler
                </a>

            </div>

            


        </form>

    </div>


    <!-- MODAL -->
    <div class="modal fade" id="cancelModal" tabindex="-1">

        <div class="modal-dialog">
            <div class="modal-content">

                <div class="modal-header">

                    <h5 class="modal-title">
                        Quitter sans enregistrer
                    </h5>

                    <button type="button"
                            class="btn-close"
                            data-bs-dismiss="modal">
                    </button>

                </div>

                <div class="modal-body">
                    Vous allez perdre toutes vos modifications.
                </div>

                <div class="modal-footer">

                    <button type="button"
                            class="btn btn-secondary"
                            data-bs-dismiss="modal">
                        Annuler
                    </button>

                    <a href="/"
                       class="btn btn-danger">
                        Quitter
                    </a>

                </div>

            </div>
        </div>

    </div>

    <!-- Bootstrap JS -->
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>

    <script>
                
        let formModified = false;

        const saveBtn = document.getElementById("saveBtn");
        const form = document.querySelector("form");

        // Détection modification formulaire
        document.querySelectorAll("input, select").forEach(function(element) {{

            element.addEventListener("input", function() {{
                formModified = true;
            }});

            element.addEventListener("change", function() {{
                formModified = true;
            }});

        }});


        // Gestion bouton Annuler
        function confirmCancel(event) {{

            // aucune modification
            if (!formModified) {{
                return true;
            }}

            // bloque navigation
            event.preventDefault();

            // ouvre modal bootstrap
            let modal = new bootstrap.Modal(
                document.getElementById('cancelModal')
            );

            modal.show();

            return false;
        }}

        
        // snapshot initial du formulaire

        function getFormState() {{

            const state = {{}};

            document.querySelector("form").querySelectorAll("input, select").forEach(el => {{

                if (el.type === "checkbox") {{

                    if (!state[el.name]) state[el.name] = [];

                    if (el.checked) {{
                        state[el.name].push(el.value);
                    }}

                }} else {{
                    state[el.name] = el.value;
                }}

            }});

            return JSON.stringify(state);
        }}

        const initialState = getFormState();

        // validation champs (rouge si vide)
        function validateFields() {{

            let valid = true;

            form.querySelectorAll("input, select").forEach(el => {{

                if (el.type !== "checkbox" && !el.value) {{
                    el.classList.add("is-invalid");
                    valid = false;
                }} else {{
                    el.classList.remove("is-invalid");
                }}

            }});

            return valid;
        }}

        // check global
        function checkForm() {{

            const currentState = getFormState();

            formModified = (currentState !== initialState);

            const valid = validateFields();

            saveBtn.disabled = !(formModified && valid);
        }}

        // écoute des changements
        form.querySelectorAll("input, select").forEach(el => {{

            el.addEventListener("input", checkForm);
            el.addEventListener("change", checkForm);

        }});

        // init
        checkForm();

        // sauvegarde
        function confirmSave(event) {{

            if (!formModified) {{

                const modal = new bootstrap.Modal(
                    document.getElementById('saveModal')
                );

                modal.show();
                return;
            }}

            form.submit();
        }}

        // MAJ affichage bouton compétences
        function updateCompetenceLabel() {{

            const checks = document.querySelectorAll(".competence-check:checked");
            const btn = document.getElementById("competenceBtn");

            if (checks.length === 0) {{
                btn.innerText = "Sélectionner vos compétences";
                return;
            }}

            let values = Array.from(checks).map(c => c.value);

            btn.innerText = values.join(", ");
        }}

        // écoute checkbox
        document.querySelectorAll(".competence-check").forEach(cb => {{
            cb.addEventListener("change", updateCompetenceLabel);
        }});


        
    </script>
    
    

    </div>
    </body>
    </html>
    """


    return HTMLResponse(content=html)




# =========================
# UPDATE GLOBAL (PAGE 1)
# =========================
@app.post("/update/{id}")
def update(
    request: Request,
    id: int,
    nom: str = Form(...),
    prenom: str = Form(...),
    profil: str = Form(...),
    agence: str = Form(...),
    competence: Optional[Union[List[str], str]] = Form(None),
    niveau: int = Form(...),
    niveau_attendu: int = Form(...),
    appetence: int = Form(...)
):

    user = get_current_user(request)

    if not user:
        raise HTTPException(status_code=401, detail="Non autorisé")

    competence = normalize_competence(competence)

    for c in collaborateurs:

        if c["id"] == id:

            c["nom"] = nom
            c["prenom"] = prenom
            c["profil"] = profil
            c["agence"] = agence
            c["competence"] = competence
            c["niveau"] = clamp(niveau)
            c["niveau_attendu"] = clamp(niveau_attendu)
            c["appetence"] = clamp(appetence)

            break

    log_action(user, "MODIFICATION", f"{prenom} {nom}")

    return RedirectResponse("/", status_code=303)


# ==========================================
# UPDATE COMPÉTENCE UNIQUE (PAGE 2 - IMPORTANT)
# ==========================================
@app.post("/update/{id}/competence/{comp_name}")
def update_competence(
    request: Request,
    id: int,
    comp_name: str,
    niveau: int = Form(...),
    niveau_attendu: int = Form(...),
    appetence: int = Form(...)
):

    user = get_current_user(request)

    if not user:
        raise HTTPException(status_code=401)

    for c in collaborateurs:
        if c["id"] == id:

            for comp in c["competence"]:
                if comp["nom"] == comp_name:

                    comp["niveau"] = clamp(niveau)
                    comp["niveau_attendu"] = clamp(niveau_attendu)
                    comp["appetence"] = clamp(appetence)
                    break
            break

    log_action(user, "MODIFICATION_COMPETENCE", comp_name)

    return RedirectResponse("/", status_code=303)



# =========================
# SAVE ALL (NOUVEAU)
# =========================

@app.post("/save-all")
def save_all(data: list = Body(...)):

    global collaborateurs

    collaborateurs = data

    return {
        "status": "success",
        "message": "Sauvegarde globale OK",
        "total": len(collaborateurs)
    }



# =========================
# SUPPRESSION COMPETENCE
# =========================

@app.get("/delete_competence/{id}/{comp_name}")
def delete_competence(
    request: Request,
    id: int,
    comp_name: str
):

    user = get_current_user(request)

    if not user:
        return RedirectResponse("/login", status_code=303)

    if not has_permission(user["role"], "delete"):
        raise HTTPException(
            status_code=403,
            detail="Suppression interdite"
        )

    c = next(
        (x for x in collaborateurs if x["id"] == id),
        None
    )

    if not c:
        return RedirectResponse("/")

    c["competence"] = normalize_competence(
        c["competence"]
    )

    c["competence"] = [
        comp
        for comp in get_safe_competences(c["competence"])
        if comp["nom"] != comp_name
    ]

    return RedirectResponse(
        f"/edit/{id}",
        status_code=303
    )




# =========================
# ANNULATION
# =========================
@app.get("/cancel")
def cancel(request: Request, action: str = "INCONNUE"):

    user = get_current_user(request)

    if not user:
        raise HTTPException(status_code=401, detail="Non autorisé")

    log_action(user, "ANNULATION", "Formulaire collaborateur abandonné")

    return RedirectResponse("/", status_code=303)




# =========================
# AUDIT LOG PAGE
# =========================
@app.get("/audit", response_class=HTMLResponse)
def audit_page():

    html = """
    <html>
    <head>
        <title>Audit Log</title>

        <link
            href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css"
            rel="stylesheet">
    </head>

    <body class="bg-light">

    <div class="container mt-4">

        <h2 class="mb-4">
            Audit Log
        </h2>

        <a href="/"
           class="btn btn-secondary mb-3">
           Retour
        </a>

        <table class="table table-bordered table-striped bg-white">

            <thead class="table-dark">
                <tr>
                    <th>Date</th>
                    <th>Utilisateur</th>
                    <th>Action</th>
                    <th>Détails</th>
                </tr>
            </thead>

            <tbody>
    """

    for log in reversed(audit_logs):

        html += f"""
        <tr>
            <td>{log['date']}</td>
            <td>{log['user']}</td>
            <td>{log['action']}</td>
            <td>{log['details']}</td>
        </tr>
        """

    html += """
            </tbody>

        </table>

    </div>

    </body>
    </html>
    """

    return HTMLResponse(content=html)





# =========================
# EXPORT EXCEL (FINAL CLEAN)
# =========================

@app.get("/export/excel")
def export_excel(
    agence: str = None,
    competence: str = None,
    profil: str = None,
    search: str = None
):

    wb = Workbook()
    ws = wb.active
    ws.title = "Skills détaillés"

    # data = collaborateurs
    data = copy.deepcopy(collaborateurs)

    if agence:
        data = [c for c in data if c["agence"] == agence]

    if profil:
        data = [c for c in data if c["profil"] == profil]

    if search:
        data = [
            c for c in data
            if search.lower() in c["nom"].lower()
            or search.lower() in c["prenom"].lower()
        ]

    if competence:
        data = [
            c for c in data
            if any(
                comp["nom"] == competence
                for comp in get_safe_competences(c["competence"])
            )
        ]

    # =========================
    # STYLES
    # =========================
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )
    center = Alignment(horizontal="center", vertical="center")

    row_fill_even = PatternFill(
        start_color="F2F2F2",
        end_color="F2F2F2",
        fill_type="solid"
    )

    # =========================
    # HEADER
    # =========================
    headers = [
        "Nom",
        "Prénom",
        "Profil",
        "Agence",
        "Compétence",
        "Niveau",
        "Niveau attendu",
        "Appétence"
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # =========================
    # DATA (1 compétence = 1 ligne)
    # =========================
    row = 2

    for c in data:

        for comp in get_safe_competences(c["competence"]):

            ws.append([
                c["nom"],
                c["prenom"],
                c["profil"],
                c["agence"],
                comp["nom"],
                comp["niveau"],
                comp["niveau_attendu"],
                comp["appetence"]
            ])

            # style ligne
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.alignment = center

                if row % 2 == 0:
                    cell.fill = row_fill_even

            row += 1

    # =========================
    # AUTO WIDTH
    # =========================
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # =========================
    # EXPORT
    # =========================
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=skills_detailles.xlsx"
        }
    )



# # =========================
# # EXPORT EXCEL (FINAL CLEAN)
# # =========================

# @app.get("/export/excel")
# def export_excel():

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Skills détaillés"

#     # =========================
#     # STYLES
#     # =========================
#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill(
#         start_color="4F81BD",
#         end_color="4F81BD",
#         fill_type="solid"
#     )
#     center = Alignment(horizontal="center", vertical="center")

#     row_fill_even = PatternFill(
#         start_color="F2F2F2",
#         end_color="F2F2F2",
#         fill_type="solid"
#     )

#     # =========================
#     # HEADER
#     # =========================
#     headers = [
#         "Nom",
#         "Prénom",
#         "Profil",
#         "Agence",
#         "Compétence",
#         "Niveau",
#         "Niveau attendu",
#         "Appétence"
#     ]

#     ws.append(headers)

#     for col in range(1, len(headers) + 1):
#         cell = ws.cell(row=1, column=col)
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = center

#     # =========================
#     # DATA (1 compétence = 1 ligne)
#     # =========================
#     row = 2

#     for c in collaborateurs:

#         for comp in get_safe_competences(c["competence"]):

#             ws.append([
#                 c["nom"],
#                 c["prenom"],
#                 c["profil"],
#                 c["agence"],
#                 comp["nom"],
#                 comp["niveau"],
#                 comp["niveau_attendu"],
#                 comp["appetence"]
#             ])

#             # style ligne
#             for col in range(1, 9):
#                 cell = ws.cell(row=row, column=col)
#                 cell.alignment = center

#                 if row % 2 == 0:
#                     cell.fill = row_fill_even

#             row += 1

#     # =========================
#     # AUTO WIDTH
#     # =========================
#     for col in ws.columns:
#         max_length = 0
#         col_letter = col[0].column_letter

#         for cell in col:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))

#         ws.column_dimensions[col_letter].width = max_length + 2

#     # =========================
#     # EXPORT
#     # =========================
#     stream = BytesIO()
#     wb.save(stream)
#     stream.seek(0)

#     return StreamingResponse(
#         stream,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={
#             "Content-Disposition": "attachment; filename=skills_detailles.xlsx"
#         }
#     )



# =====================
# EXPORT JSON
# =====================
@app.get("/export/json")
def export_json():
    return {"count": len(collaborateurs), "data": collaborateurs}




# =========================
# EXPORT PDF RH OFFICIEL
# =========================

@app.get("/export/pdf")
def export_pdf(
    agence: str = None,
    competence: str = None,
    profil: str = None,
    search: str = None
):


    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # data = collaborateurs
    data = copy.deepcopy(collaborateurs)

    if agence:
        data = [c for c in data if c["agence"] == agence]

    if profil:
        data = [c for c in data if c["profil"] == profil]

    if search:
        data = [
            c for c in data
            if search.lower() in c["nom"].lower()
            or search.lower() in c["prenom"].lower()
        ]

    if competence:
        data = [
            c for c in data
            if any(
                comp["nom"] == competence
                for comp in get_safe_competences(c["competence"])
            )
        ]


    # =========================
    # TITRE
    # =========================

    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(40, height - 35, "Rapport RH - Skills Matrix")

    pdf.setFont("Helvetica", 10)
    pdf.drawString(
        40,
        height - 50,
        f"Nombre de collaborateurs : {len(data)}"
    )

    # =========================
    # HEADER
    # =========================

    def draw_header():

        pdf.setFillColorRGB(0.31, 0.51, 0.74)

        pdf.rect(
            40,
            height - 90,
            520,
            20,
            fill=1
        )

        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 9)

        pdf.drawString(45,  height - 85, "Nom")
        pdf.drawString(115, height - 85, "Prénom")
        pdf.drawString(185, height - 85, "Profil")
        pdf.drawString(285, height - 85, "Agence")
        pdf.drawString(355, height - 85, "Compétence")
        pdf.drawString(445, height - 85, "Niv")
        pdf.drawString(485, height - 85, "Att")
        pdf.drawString(525, height - 85, "App")

    draw_header()

    y = height - 110

    # =========================
    # DATA
    # =========================

    for c in data:

        competences = get_safe_competences(c["competence"])

        score = 0

        if competences:
            score = round(
                sum(comp["niveau"] for comp in competences)
                / len(competences),
                1
            )

        for comp in competences:

            if y < 80:

                pdf.showPage()

                draw_header()

                y = height - 110

            # =========================
            # LIGNE
            # =========================

            pdf.setStrokeColor(colors.black)

            pdf.rect(
                40,
                y - 5,
                520,
                15,
                stroke=1,
                fill=0
            )

            pdf.setFillColor(colors.black)
            pdf.setFont("Helvetica", 8)

            pdf.drawString(45, y, str(c["nom"]))
            pdf.drawString(115, y, str(c["prenom"]))
            pdf.drawString(185, y, str(c["profil"]))
            pdf.drawString(285, y, str(c["agence"]))
            pdf.drawString(355, y, str(comp["nom"]))

            # =========================
            # BARRE NIVEAU
            # =========================

            niveau = comp["niveau"]

            bar_x = 445
            bar_y = y
            bar_width = 30
            bar_height = 6

            pdf.setFillColor(colors.lightgrey)

            pdf.rect(
                bar_x,
                bar_y,
                bar_width,
                bar_height,
                fill=1,
                stroke=0
            )

            ratio = min(niveau / 5, 1)

            pdf.setFillColor(colors.green)

            pdf.rect(
                bar_x,
                bar_y,
                bar_width * ratio,
                bar_height,
                fill=1,
                stroke=0
            )

            pdf.setFillColor(colors.black)

            pdf.drawString(
                485,
                y,
                str(comp["niveau_attendu"])
            )

            pdf.drawString(
                525,
                y,
                str(comp["appetence"])
            )

            y -= 15

        # =========================
        # SCORE
        # =========================

        if y < 80:

            pdf.showPage()

            draw_header()

            y = height - 110

        if score >= 4:
            pdf.setFillColor(colors.darkgreen)

        elif score >= 3:
            pdf.setFillColor(colors.orange)

        else:
            pdf.setFillColor(colors.red)

        pdf.setFont("Helvetica-Bold", 9)

        pdf.drawString(
            45,
            y,
            f"Score global {c['nom']} {c['prenom']} : {score}/5"
        )

        pdf.setFillColor(colors.black)

        y -= 25

    pdf.showPage()

    pdf.setFont("Helvetica-Bold", 18)

    pdf.drawString(
        40,
        height - 60,
        "Synthèse RH"
    )

    pdf.setFont("Helvetica", 12)

    pdf.drawString(
        40,
        height - 100,
        f"Nombre de collaborateurs : {len(data)}"
    )

    total_competences = sum(
        len(get_safe_competences(c["competence"]))
        for c in data
    )

    pdf.drawString(
        40,
        height - 130,
        f"Nombre total de compétences : {total_competences}"
    )

    pdf.save()

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "attachment; filename=rapport_RH_officiel.pdf"
        }
    )





# # =====================================
# # EXPORT PDF RH OFFICIEL - MULTI PAGES
# # =====================================


# @app.get("/export/pdf")

# def export_pdf():

#     buffer = BytesIO()
#     pdf = canvas.Canvas(buffer, pagesize=A4)
#     width, height = A4

#     # =========================
#     # PAGE 1 : DASHBOARD
#     # =========================

#     pdf.setFont("Helvetica-Bold", 14)
#     pdf.drawString(40, height - 50, "Dashboard RH")

#     top_skills = build_top_skills(collaborateurs)
#     chart_path = draw_top_skills_chart(top_skills)

#     pdf.drawImage(chart_path, 40, height - 300, width=400, height=200)

#     pdf.showPage()

#     # =========================
#     # PAGES COLLABORATEURS
#     # =========================

#     for c in collaborateurs:

#         pdf.setFont("Helvetica-Bold", 12)
#         pdf.drawString(40, height - 50, f"{c['nom']} {c['prenom']}")

#         score = compute_score(c)

#         pdf.setFont("Helvetica", 10)
#         pdf.drawString(40, height - 80, f"Score intelligent: {score}")

#         radar_path = draw_radar(c)

#         if radar_path:
#             pdf.drawImage(radar_path, 40, height - 300, width=300, height=200)

#         # SKILL MATRIX
#         y = height - 330

#         pdf.setFont("Helvetica-Bold", 10)
#         pdf.drawString(40, y, "Compétence | Niveau | Attendu")
#         y -= 20

#         pdf.setFont("Helvetica", 9)

#         for comp in get_safe_competences(c["competence"]):

#             pdf.drawString(
#                 40,
#                 y,
#                 f"{comp['nom']} | {comp['niveau']} | {comp['niveau_attendu']}"
#             )
#             y -= 15

#         pdf.showPage()

#     pdf.save()
#     buffer.seek(0)

#     return StreamingResponse(
#         buffer,
#         media_type="application/pdf",
#         headers={"Content-Disposition": "attachment; filename=RH_dashboard.pdf"}
#     )


# # =========================
# # FONCTION TOP SKILLS
# # =========================

# def build_top_skills(data):
#     skills = []

#     for c in data:
#         for comp in get_safe_competences(c["competence"]):
#             skills.append(comp["nom"])

#     counter = Counter(skills)
#     return counter.most_common(5)



# # =========================
# # FONCTION GRAPH TOP SKILLS
# # =========================

# def draw_top_skills_chart(top_skills):
    
#     labels = [skill for skill, count in top_skills]
#     values = [count for skill, count in top_skills]

#     plt.figure(figsize=(8, 4))

#     plt.bar(labels, values)

#     plt.title("Top 5 compétences")
#     plt.ylabel("Nombre de collaborateurs")

#     path = os.path.join(
#         tempfile.gettempdir(),
#         "top_skills.png"
#     )

#     plt.tight_layout()
#     plt.savefig(path)
#     plt.close()

#     return path



# # =========================================
# # FONCTION RADAR CHART (par collaborateur)
# # =========================================

# def draw_radar(c):
    
#     comps = get_safe_competences(c["competence"])

#     labels = [x["nom"] for x in comps]
#     values = [x["niveau"] for x in comps]

#     if not labels:
#         return None

#     angles = np.linspace(
#         0,
#         2 * np.pi,
#         len(labels),
#         endpoint=False
#     ).tolist()

#     values += values[:1]
#     angles += angles[:1]

#     fig, ax = plt.subplots(
#         subplot_kw=dict(polar=True)
#     )

#     ax.plot(angles, values)
#     ax.fill(angles, values, alpha=0.3)

#     ax.set_xticks(angles[:-1])
#     ax.set_xticklabels(labels)

#     path = os.path.join(
#         tempfile.gettempdir(),
#         f"radar_{c['id']}.png"
#     )

#     plt.savefig(path)
#     plt.close()

#     return path



# # =============================
# # FONCTION SCORING INTELLIGENT
# # =============================

# def compute_score(c):
    
#     comps = get_safe_competences(c["competence"])

#     if not comps:
#         return 0

#     return round(
#         sum(comp["niveau"] / comp["niveau_attendu"] if comp["niveau_attendu"] else 0
#             for comp in comps)
#         / len(comps),
#         2
#     )













# # =========================
# # EXPORT PDF RH OFFICIEL
# # =========================

# @app.get("/export/pdf")
# def export_pdf():

#     buffer = BytesIO()
#     pdf = canvas.Canvas(buffer, pagesize=A4)

#     width, height = A4
    


#     # =========================
#     # HEADER FIXE (fonction)
#     # =========================
#     def draw_header():
#         pdf.setFont("Helvetica-Bold", 10)
#         pdf.setFillColor(colors.white)

#         pdf.setFillColorRGB(0.31, 0.51, 0.74)
#         pdf.rect(40, height - 60, 520, 20, fill=1)

#         pdf.setFillColor(colors.white)

#         pdf.drawString(45, height - 55, "Nom")
#         pdf.drawString(115, height - 55, "Prénom")
#         pdf.drawString(185, height - 55, "Profil")
#         pdf.drawString(285, height - 55, "Agence")
#         pdf.drawString(365, height - 55, "Compétence")
#         pdf.drawString(450, height - 55, "Niv")
#         pdf.drawString(480, height - 55, "Att")
#         pdf.drawString(510, height - 55, "App")

#     # =========================
#     # INIT
#     # =========================
#     y = height - 80
#     pdf.setFont("Helvetica", 9)

#     draw_header()

#     # =========================
#     # DATA
#     # =========================
#     for c in collaborateurs:

#         competences = get_safe_competences(c["competence"])

#         # 🧠 SCORE GLOBAL collaborateur
#         score = 0
#         if competences:
#             score = sum(comp["niveau"] for comp in competences) / len(competences)

#         for comp in competences:

#             if y < 80:
#                 pdf.showPage()
#                 y = height - 80
#                 draw_header()

#             # =========================
#             # BORDURE LIGNE
#             # =========================
#             pdf.setStrokeColor(colors.black)
#             pdf.rect(40, y - 5, 520, 15, stroke=1, fill=0)

#             # =========================
#             # TEXT
#             # =========================
#             pdf.setFillColor(colors.black)

#             pdf.drawString(45, y, str(c["nom"]))
#             pdf.drawString(115, y, str(c["prenom"]))
#             pdf.drawString(185, y, str(c["profil"]))
#             pdf.drawString(285, y, str(c["agence"]))
#             pdf.drawString(365, y, str(comp["nom"]))

#             # =========================
#             # BARRE NIVEAU VISUEL
#             # =========================
#             niveau = comp["niveau"]

#             bar_x = 450
#             bar_y = y
#             bar_width = 50
#             bar_height = 6

#             # fond barre
#             pdf.setFillColor(colors.lightgrey)
#             pdf.rect(bar_x, bar_y, bar_width, bar_height, fill=1, stroke=0)

#             # progression
#             ratio = min(niveau / 5, 1)
#             pdf.setFillColor(colors.green)
#             pdf.rect(bar_x, bar_y, bar_width * ratio, bar_height, fill=1, stroke=0)

#             pdf.setFillColor(colors.black)

#             pdf.drawString(510, y, str(comp["niveau_attendu"]))

#             y -= 15

#         # =========================
#         # SCORE COLLABORATEUR
#         # =========================
#         pdf.setFont("Helvetica-Bold", 9)
#         pdf.drawString(45, y, f"Score global {c['nom']} {c['prenom']} : {round(score,1)} / 5")
#         pdf.setFont("Helvetica", 9)

#         y -= 25

#     pdf.save()
#     buffer.seek(0)

#     return StreamingResponse(
#         buffer,
#         media_type="application/pdf",
#         headers={
#             "Content-Disposition": "attachment; filename=rapport_RH_officiel.pdf"
#         }
#     )



# =========================
# EXPORT PDF RH OFFICIEL
# =========================


# @app.get("/export/pdf")
# def export_pdf():

#     buffer = BytesIO()
#     pdf = canvas.Canvas(buffer, pagesize=A4)

#     width, height = A4
#     y = height - 50

#     # =========================
#     # HEADER
#     # =========================
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.setFillColor(colors.white)

#     pdf.setFillColorRGB(0.31, 0.51, 0.74)  # bleu header

#     pdf.rect(40, y - 10, 520, 20, fill=1)

#     pdf.setFillColor(colors.white)

#     pdf.drawString(45, y, "Nom")
#     pdf.drawString(115, y, "Prénom")
#     pdf.drawString(185, y, "Profil")
#     pdf.drawString(285, y, "Agence")
#     pdf.drawString(365, y, "Compétence")
#     pdf.drawString(450, y, "Niv")
#     pdf.drawString(480, y, "Att")
#     pdf.drawString(510, y, "App")

#     y -= 25

#     pdf.setFont("Helvetica", 9)
#     pdf.setFillColor(colors.black)

#     # =========================
#     # DATA
#     # =========================
#     for c in collaborateurs:

#         for comp in get_safe_competences(c["competence"]):

#             if y < 60:
#                 pdf.showPage()
#                 y = height - 50
#                 pdf.setFont("Helvetica", 9)

#             # =========================
#             # CELL BORDER (ligne encadrée)
#             # =========================
#             pdf.setStrokeColor(colors.black)
#             pdf.setLineWidth(0.5)

#             pdf.rect(40, y - 5, 520, 15, stroke=1, fill=0)

#             # =========================
#             # TEXT
#             # =========================
#             pdf.drawString(45, y, str(c["nom"]))
#             pdf.drawString(115, y, str(c["prenom"]))
#             pdf.drawString(185, y, str(c["profil"]))
#             pdf.drawString(285, y, str(c["agence"]))
#             pdf.drawString(365, y, str(comp["nom"]))

#             # couleur niveau
#             niveau = comp["niveau"]

#             if niveau <= 1:
#                 pdf.setFillColor(colors.red)
#             elif niveau <= 3:
#                 pdf.setFillColor(colors.orange)
#             else:
#                 pdf.setFillColor(colors.green)

#             pdf.drawString(450, y, str(comp["niveau"]))

#             pdf.setFillColor(colors.black)

#             pdf.drawString(480, y, str(comp["niveau_attendu"]))
#             pdf.drawString(510, y, str(comp["appetence"]))

#             y -= 15

#     pdf.save()
#     buffer.seek(0)

#     return StreamingResponse(
#         buffer,
#         media_type="application/pdf",
#         headers={
#             "Content-Disposition": "attachment; filename=skills_table_encadre.pdf"
#         }
#     )